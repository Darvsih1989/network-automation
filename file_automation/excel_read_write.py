from openpyxl import Workbook, load_workbook

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "TestSheet"

# Write data
ws['A1'] = "Name"
ws['B1'] = "Value"
ws.append(["Alice", 100])
ws.append(["Bob", 200])

# Save workbook
wb.save("example.xlsx")
print("Excel file saved as example.xlsx")

# Read data
wb2 = load_workbook("example.xlsx")
ws2 = wb2["TestSheet"]
for row in ws2.iter_rows(values_only=True):
    print(row)